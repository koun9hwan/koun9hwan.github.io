import pandas as pd     # pip install pands필요
import numpy as np
from openpyxl import load_workbook



# 해당이름의 csv파일을 읽어옴
r_csv = pd.read_csv("진단결과.csv")

# 저장할 xlsx파일의 이름을 정함
save_xlsx = pd.ExcelWriter("진단결과.xlsx")

r_csv.to_excel(save_xlsx, index = False) # xlsx 파일로 변환

save_xlsx.save() #xlsx 파일로 저장

old_wb=load_workbook("진단결과.xlsx")
old_ws=old_wb.active

new_wb=load_workbook("주요정보통신기반시설_취약점_진단_가이드_보고서.xlsx")
new_ws=new_wb.get_sheet_by_name("스크립트 결과 값 입력 시트")

for i in range(1, 74):
    for j in range(1, 4):
        old_cell=old_ws.cell(i,j)
        new_cell=new_ws.cell(i,j, value=old_cell.value)

new_wb.save("주요정보통신기반시설_취약점_진단_가이드_보고서.xlsx")



